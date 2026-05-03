"""
设计项目进度管理系统 - Flask后端
技术栈: Flask + SQLite + Flask-Login
"""
import os
import sys
import io
from datetime import datetime, timedelta
from functools import wraps

from flask import Flask, request, jsonify, send_from_directory, session, send_file, make_response
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook, load_workbook

try:
    from pypinyin import pinyin, Style
    USE_PYPINYIN = True
except ImportError:
    USE_PYPINYIN = False

# 简化的拼音映射（当pypinyin不可用时）
SIMPLE_PINYIN = {
    '王': 'wang', '李': 'li', '张': 'zhang', '刘': 'liu', '陈': 'chen',
    '杨': 'yang', '黄': 'huang', '赵': 'zhao', '周': 'zhou', '吴': 'wu',
    '徐': 'xu', '孙': 'sun', '马': 'ma', '朱': 'zhu', '胡': 'hu',
    '林': 'lin', '何': 'he', '郭': 'guo', '罗': 'luo', '高': 'gao',
    '郑': 'zheng', '梁': 'liang', '谢': 'xie', '宋': 'song', '唐': 'tang',
    '许': 'xu', '韩': 'han', '冯': 'feng', '邓': 'deng', '曹': 'cao',
    '彭': 'peng', '曾': 'zeng', '肖': 'xiao', '田': 'tian', '董': 'dong',
    '袁': 'yuan', '潘': 'pan', '于': 'yu', '蒋': 'jiang', '蔡': 'cai',
    '余': 'yu', '杜': 'du', '叶': 'ye', '程': 'cheng', '苏': 'su',
    '魏': 'wei', '吕': 'lv', '丁': 'ding', '任': 'ren', '沈': 'shen',
    '姚': 'yao', '卢': 'lu', '姜': 'jiang', '崔': 'cui', '钟': 'zhong',
    '谭': 'tan', '陆': 'lu', '汪': 'wang', '范': 'fan', '金': 'jin',
    '石': 'shi', '廖': 'liao', '贾': 'jia', '夏': 'xia', '韦': 'wei',
    '傅': 'fu', '方': 'fang', '白': 'bai', '邹': 'zou', '孟': 'meng',
    '熊': 'xiong', '秦': 'qin', '邱': 'qiu', '侯': 'hou', '江': 'jiang',
    '尹': 'yin', '薛': 'xue', '闫': 'yan', '段': 'duan', '雷': 'lei',
    '龙': 'long', '史': 'shi', '陶': 'tao', '贺': 'he', '顾': 'gu',
    '毛': 'mao', '郝': 'hao', '龚': 'gong', '邵': 'shao', '万': 'wan',
    '钱': 'qian', '严': 'yan', '孔': 'kong', '常': 'chang', '武': 'wu',
    '乔': 'qiao', '赖': 'lai', '庞': 'pang', '樊': 'fan', '殷': 'yin',
    '施': 'shi', '陶': 'tao', '洪': 'hong', '翟': 'zhai', '安': 'an'
}


def generate_username(chinese_name):
    """根据中文姓名生成用户名"""
    if not chinese_name or not isinstance(chinese_name, str):
        return 'user'
    
    name = chinese_name.strip()
    if len(name) == 0:
        return 'user'
    
    pinyin_result = []
    
    if USE_PYPINYIN:
        # 使用pypinyin库
        try:
            pinyin_result = [p[0].lower() for p in pinyin(name, style=Style.NORMAL)]
        except:
            pass
    
    if not pinyin_result:
        # 使用简化拼音映射
        pinyin_result = []
        for char in name:
            pinyin_result.append(SIMPLE_PINYIN.get(char, char.lower()))
    
    if len(pinyin_result) == 1:
        # 单字姓名
        return pinyin_result[0]
    elif len(pinyin_result) == 2:
        # 双字姓名：全拼
        return ''.join(pinyin_result)
    else:
        # 三字及以上：姓全拼 + 名首字母
        surname = pinyin_result[0]
        given_name_initials = ''.join([p[0] for p in pinyin_result[1:]])
        return surname + given_name_initials

# ==================== App Config ====================
app = Flask(__name__, static_folder='static', template_folder='templates')
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dpm-secret-key-change-in-production-2026')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///design_manager.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

CORS(app, supports_credentials=True)
db = SQLAlchemy(app)

# ==================== Database Models ====================

class User(db.Model):
    """用户表"""
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    real_name = db.Column(db.String(80), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='staff')  # admin, engineer, staff
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.now)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def to_dict(self):
        return {
            'id': self.id,
            'username': self.username,
            'realName': self.real_name,
            'role': self.role,
            'isActive': self.is_active,
            'createdAt': self.created_at.strftime('%Y-%m-%d %H:%M:%S') if self.created_at else ''
        }


class Designer(db.Model):
    """设计人员表"""
    __tablename__ = 'designers'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False)  # 姓名
    department = db.Column(db.String(100))  # 所属部门
    major = db.Column(db.String(50))  # 专业（建筑/结构/给排水/暖通/电气等）
    title = db.Column(db.String(50))  # 职称
    phone = db.Column(db.String(20))  # 联系电话
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.now)

    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'department': self.department,
            'major': self.major,
            'title': self.title,
            'phone': self.phone,
            'isActive': self.is_active,
            'createdAt': self.created_at.strftime('%Y-%m-%d %H:%M:%S') if self.created_at else ''
        }


class ConstructionPlan(db.Model):
    """施工图计划表"""
    __tablename__ = 'construction_plans'
    id = db.Column(db.Integer, primary_key=True)
    project_name = db.Column(db.String(200), nullable=False)
    unit_project = db.Column(db.String(200))
    drawing_content = db.Column(db.String(500))
    major_category = db.Column(db.String(50))
    designer = db.Column(db.String(50))
    completion_status = db.Column(db.String(20))
    start_date = db.Column(db.String(20))
    end_date = db.Column(db.String(20))
    project_manager = db.Column(db.String(50))
    remarks = db.Column(db.String(500))
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

    def to_dict(self):
        return {
            'id': self.id,
            'projectName': self.project_name,
            'unitProject': self.unit_project,
            'drawingContent': self.drawing_content,
            'majorCategory': self.major_category,
            'designer': self.designer,
            'completionStatus': self.completion_status,
            'startDate': self.start_date,
            'endDate': self.end_date,
            'projectManager': self.project_manager,
            'remarks': self.remarks,
            'updatedAt': self.updated_at.strftime('%Y-%m-%d %H:%M:%S') if self.updated_at else ''
        }


class PhasePlan(db.Model):
    """阶段设计计划表"""
    __tablename__ = 'phase_plans'
    id = db.Column(db.Integer, primary_key=True)
    project_name = db.Column(db.String(200), nullable=False)
    design_phase = db.Column(db.String(50))
    project_manager = db.Column(db.String(50))
    department = db.Column(db.String(100))
    participating_majors = db.Column(db.String(200))
    data_submit_date = db.Column(db.String(20))
    desc_submit_date = db.Column(db.String(20))
    publish_date = db.Column(db.String(20))
    remarks = db.Column(db.String(500))
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

    def to_dict(self):
        return {
            'id': self.id,
            'projectName': self.project_name,
            'designPhase': self.design_phase,
            'projectManager': self.project_manager,
            'department': self.department,
            'participatingMajors': self.participating_majors,
            'dataSubmitDate': self.data_submit_date,
            'descSubmitDate': self.desc_submit_date,
            'publishDate': self.publish_date,
            'remarks': self.remarks,
            'updatedAt': self.updated_at.strftime('%Y-%m-%d %H:%M:%S') if self.updated_at else ''
        }


class TechnicalPlan(db.Model):
    """技术要求计划表"""
    __tablename__ = 'technical_plans'
    id = db.Column(db.Integer, primary_key=True)
    equipment_name = db.Column(db.String(200), nullable=False)
    designer = db.Column(db.String(50))
    design_complete_date = db.Column(db.String(20))
    purchase_date = db.Column(db.String(20))
    owner_review_date = db.Column(db.String(20))
    owner_feedback_date = db.Column(db.String(20))
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

    def to_dict(self):
        return {
            'id': self.id,
            'equipmentName': self.equipment_name,
            'designer': self.designer,
            'designCompleteDate': self.design_complete_date,
            'purchaseDate': self.purchase_date,
            'ownerReviewDate': self.owner_review_date,
            'ownerFeedbackDate': self.owner_feedback_date,
            'updatedAt': self.updated_at.strftime('%Y-%m-%d %H:%M:%S') if self.updated_at else ''
        }


class MajorCategory(db.Model):
    """专业类别表"""
    __tablename__ = 'major_categories'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False, unique=True)
    created_at = db.Column(db.DateTime, default=datetime.now)

    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'createdAt': self.created_at.strftime('%Y-%m-%d %H:%M:%S') if self.created_at else ''
        }


# ==================== Auth Helpers ====================

def login_required(f):
    """登录验证装饰器"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'code': 401, 'msg': '请先登录'}), 401
        return f(*args, **kwargs)
    return decorated


def edit_required(f):
    """编辑权限验证装饰器（admin或engineer）"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'code': 401, 'msg': '请先登录'}), 401
        user = User.query.get(session['user_id'])
        if not user or user.role == 'staff':
            return jsonify({'code': 403, 'msg': '权限不足，仅首席工程师及以上可操作'}), 403
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    """管理员权限验证装饰器"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'code': 401, 'msg': '请先登录'}), 401
        user = User.query.get(session['user_id'])
        if not user or user.role != 'admin':
            return jsonify({'code': 403, 'msg': '仅管理员可操作'}), 403
        return f(*args, **kwargs)
    return decorated


def get_current_user():
    """获取当前登录用户"""
    if 'user_id' in session:
        return User.query.get(session['user_id'])
    return None


# ==================== Init Default Data ====================

def init_default_data():
    """初始化默认用户和示例数据"""
    # Create default users
    defaults = [
        {'username': 'admin', 'password': 'admin123', 'real_name': '系统管理员', 'role': 'admin'},
        {'username': 'engineer', 'password': 'eng123', 'real_name': '张工（首席工程师）', 'role': 'engineer'},
        {'username': 'staff', 'password': 'staff123', 'real_name': '李明（设计人员）', 'role': 'staff'},
    ]
    for d in defaults:
        if not User.query.filter_by(username=d['username']).first():
            user = User(username=d['username'], real_name=d['real_name'], role=d['role'])
            user.set_password(d['password'])
            db.session.add(user)
    db.session.commit()

    # Create default designers
    if Designer.query.count() == 0:
        default_designers = [
            Designer(name='王工', department='设计一部', major='建筑', title='工程师'),
            Designer(name='李工', department='设计一部', major='结构', title='工程师'),
            Designer(name='赵工', department='设计一部', major='给排水', title='高级工程师'),
            Designer(name='陈工', department='设计二部', major='暖通', title='工程师'),
            Designer(name='孙工', department='设计二部', major='电气', title='工程师'),
            Designer(name='刘工', department='设计二部', major='总图', title='高级工程师'),
            Designer(name='张工', department='设计一部', major='建筑', title='高级工程师'),
        ]
        db.session.add_all(default_designers)
        db.session.commit()

    # Create default major categories
    if MajorCategory.query.count() == 0:
        default_majors = [
            MajorCategory(name='建筑'),
            MajorCategory(name='结构'),
            MajorCategory(name='给排水'),
            MajorCategory(name='暖通'),
            MajorCategory(name='电气'),
            MajorCategory(name='总图'),
            MajorCategory(name='动力'),
            MajorCategory(name='通信'),
            MajorCategory(name='其他'),
        ]
        db.session.add_all(default_majors)
        db.session.commit()

    # Create sample data for construction plans
    if ConstructionPlan.query.count() == 0:
        samples = [
            ConstructionPlan(project_name='XX污水处理厂扩建工程', unit_project='生化处理车间',
                           drawing_content='建筑平立剖面图', major_category='建筑', designer='王工',
                           completion_status='进行中', start_date='2026-03-01', end_date='2026-06-30',
                           project_manager='张工', remarks='一期扩建'),
            ConstructionPlan(project_name='XX污水处理厂扩建工程', unit_project='生化处理车间',
                           drawing_content='结构施工图', major_category='结构', designer='李工',
                           completion_status='未开始', start_date='2026-04-01', end_date='2026-07-15',
                           project_manager='张工', remarks=''),
            ConstructionPlan(project_name='XX污水处理厂扩建工程', unit_project='生化处理车间',
                           drawing_content='给排水管线图', major_category='给排水', designer='赵工',
                           completion_status='已完成', start_date='2026-02-15', end_date='2026-05-01',
                           project_manager='张工', remarks='已完成审查'),
            ConstructionPlan(project_name='YY热电厂改造项目', unit_project='主厂房',
                           drawing_content='暖通空调施工图', major_category='暖通', designer='陈工',
                           completion_status='进行中', start_date='2026-03-15', end_date='2026-08-01',
                           project_manager='刘工', remarks='含洁净室设计'),
            ConstructionPlan(project_name='YY热电厂改造项目', unit_project='主厂房',
                           drawing_content='电气施工图', major_category='电气', designer='孙工',
                           completion_status='已延期', start_date='2026-03-01', end_date='2026-06-15',
                           project_manager='刘工', remarks='设备资料延迟'),
        ]
        db.session.add_all(samples)

    # Create sample data for phase plans
    if PhasePlan.query.count() == 0:
        samples = [
            PhasePlan(project_name='XX污水处理厂扩建工程', design_phase='初步设计',
                     project_manager='张工', department='设计一部',
                     participating_majors='建筑、结构、给排水',
                     data_submit_date='2026-02-01', desc_submit_date='2026-03-15',
                     publish_date='2026-04-30', remarks='已通过评审'),
            PhasePlan(project_name='XX污水处理厂扩建工程', design_phase='施工图设计',
                     project_manager='张工', department='设计一部',
                     participating_majors='建筑、结构、给排水、暖通、电气',
                     data_submit_date='2026-05-01', desc_submit_date='', publish_date='',
                     remarks='进行中'),
            PhasePlan(project_name='YY热电厂改造项目', design_phase='方案设计',
                     project_manager='刘工', department='设计二部',
                     participating_majors='建筑、结构、电气',
                     data_submit_date='2026-01-15', desc_submit_date='2026-02-28',
                     publish_date='2026-03-15', remarks='方案已确认'),
            PhasePlan(project_name='ZZ工业园区基础设施', design_phase='初步设计',
                     project_manager='王工', department='设计三部',
                     participating_majors='总图、建筑、结构、给排水',
                     data_submit_date='2026-04-01', desc_submit_date='', publish_date='',
                     remarks='新项目'),
        ]
        db.session.add_all(samples)

    # Create sample data for technical plans
    if TechnicalPlan.query.count() == 0:
        samples = [
            TechnicalPlan(equipment_name='离心式鼓风机（Q=120m3/min）', designer='赵工',
                         design_complete_date='2026-03-15', purchase_date='2026-04-01',
                         owner_review_date='2026-04-15', owner_feedback_date='2026-05-01'),
            TechnicalPlan(equipment_name='潜水排污泵（Q=200m3/h）', designer='赵工',
                         design_complete_date='2026-03-20', purchase_date='2026-04-10',
                         owner_review_date='2026-04-20', owner_feedback_date=''),
            TechnicalPlan(equipment_name='板式换热器（F=50m2）', designer='陈工',
                         design_complete_date='2026-04-01', purchase_date='',
                         owner_review_date='', owner_feedback_date=''),
            TechnicalPlan(equipment_name='PLC控制柜', designer='孙工',
                         design_complete_date='2026-03-10', purchase_date='2026-03-25',
                         owner_review_date='2026-04-05', owner_feedback_date='2026-04-20'),
        ]
        db.session.add_all(samples)

    db.session.commit()


# ==================== Page Routes ====================

@app.route('/')
def index():
    """返回主页面"""
    resp = make_response(send_from_directory('.', 'index.html'))
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp


# ==================== Auth API ====================

@app.route('/api/login', methods=['POST'])
def api_login():
    """用户登录"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    username = data.get('username', '').strip()
    password = data.get('password', '')

    if not username or not password:
        return jsonify({'code': 400, 'msg': '用户名和密码不能为空'}), 400

    user = User.query.filter_by(username=username, is_active=True).first()
    if not user or not user.check_password(password):
        return jsonify({'code': 401, 'msg': '用户名或密码错误'}), 401

    session['user_id'] = user.id
    session['username'] = user.username

    return jsonify({
        'code': 200,
        'msg': '登录成功',
        'data': user.to_dict()
    })


@app.route('/api/logout', methods=['POST'])
def api_logout():
    """用户登出"""
    session.clear()
    return jsonify({'code': 200, 'msg': '已退出登录'})


@app.route('/api/current-user', methods=['GET'])
def api_current_user():
    """获取当前登录用户信息"""
    user = get_current_user()
    if not user:
        return jsonify({'code': 401, 'msg': '未登录'}), 401
    return jsonify({'code': 200, 'data': user.to_dict()})


# ==================== User Management API ====================

@app.route('/api/users', methods=['GET'])
@login_required
def api_get_users():
    """获取用户列表（仅管理员）"""
    user = get_current_user()
    if user.role != 'admin':
        return jsonify({'code': 403, 'msg': '仅管理员可查看用户列表'}), 403

    users = User.query.order_by(User.id).all()
    return jsonify({
        'code': 200,
        'data': [u.to_dict() for u in users]
    })


@app.route('/api/users', methods=['POST'])
@admin_required
def api_create_user():
    """创建用户"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    real_name = data.get('realName', '').strip()
    role = data.get('role', 'staff')

    if not username or not password or not real_name:
        return jsonify({'code': 400, 'msg': '用户名、密码和姓名不能为空'}), 400

    if role not in ('admin', 'engineer', 'staff'):
        return jsonify({'code': 400, 'msg': '无效的角色类型'}), 400

    if User.query.filter_by(username=username).first():
        return jsonify({'code': 400, 'msg': '用户名已存在'}), 400

    user = User(username=username, real_name=real_name, role=role)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()

    return jsonify({'code': 200, 'msg': '用户创建成功', 'data': user.to_dict()})


@app.route('/api/users/<int:user_id>', methods=['PUT'])
@admin_required
def api_update_user(user_id):
    """更新用户"""
    user = User.query.get(user_id)
    if not user:
        return jsonify({'code': 404, 'msg': '用户不存在'}), 404

    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    if 'realName' in data:
        user.real_name = data['realName'].strip()
    if 'role' in data and data['role'] in ('admin', 'engineer', 'staff'):
        user.role = data['role']
    if 'isActive' in data:
        user.is_active = data['isActive']
    if 'password' in data and data['password'].strip():
        user.set_password(data['password'].strip())

    db.session.commit()
    return jsonify({'code': 200, 'msg': '用户更新成功', 'data': user.to_dict()})


@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@admin_required
def api_delete_user(user_id):
    """删除用户"""
    if user_id == session.get('user_id'):
        return jsonify({'code': 400, 'msg': '不能删除当前登录的用户'}), 400

    user = User.query.get(user_id)
    if not user:
        return jsonify({'code': 404, 'msg': '用户不存在'}), 404

    db.session.delete(user)
    db.session.commit()
    return jsonify({'code': 200, 'msg': '用户已删除'})


@app.route('/api/users/sync-from-designers', methods=['POST'])
@admin_required
def api_sync_users_from_designers():
    """从设计人员同步用户（自动创建启用状态的设计人员为用户）"""
    # 获取所有启用状态的设计人员
    active_designers = Designer.query.filter_by(is_active=True).all()
    
    # 获取现有用户名
    existing_usernames = {u.username for u in User.query.all()}
    
    created_count = 0
    skipped_count = 0
    
    for designer in active_designers:
        # 生成用户名
        username = generate_username(designer.name)
        
        # 如果用户名已存在，尝试添加数字后缀
        original_username = username
        suffix = 1
        while username in existing_usernames:
            username = f"{original_username}{suffix}"
            suffix += 1
        
        # 检查是否已存在该用户
        if User.query.filter_by(real_name=designer.name).first():
            skipped_count += 1
            continue
        
        # 创建用户，密码为用户名+123
        password = f"{username}123"
        user = User(
            username=username,
            real_name=designer.name,
            role='engineer'  # 设计人员身份为设计人员角色
        )
        user.set_password(password)
        db.session.add(user)
        existing_usernames.add(username)
        created_count += 1
    
    db.session.commit()
    
    return jsonify({
        'code': 200,
        'msg': f'同步完成！新增 {created_count} 个用户，跳过 {skipped_count} 个已有用户'
    })


@app.route('/api/users/change-password', methods=['POST'])
@login_required
def api_change_password():
    """用户修改密码"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400
    
    old_password = data.get('oldPassword', '').strip()
    new_password = data.get('newPassword', '').strip()
    
    if not old_password or not new_password:
        return jsonify({'code': 400, 'msg': '请输入原密码和新密码'}), 400
    
    if len(new_password) < 6:
        return jsonify({'code': 400, 'msg': '新密码长度至少为6位'}), 400
    
    user = User.query.get(session.get('user_id'))
    if not user:
        return jsonify({'code': 404, 'msg': '用户不存在'}), 404
    
    if not user.check_password(old_password):
        return jsonify({'code': 400, 'msg': '原密码不正确'}), 400
    
    user.set_password(new_password)
    db.session.commit()
    
    return jsonify({'code': 200, 'msg': '密码修改成功'})


# ==================== Designer API ====================

@app.route('/api/designers', methods=['GET'])
@login_required
def api_get_designers():
    """获取所有设计人员列表（支持keyword搜索name）"""
    keyword = request.args.get('keyword', '').strip()
    query = Designer.query.order_by(Designer.id)
    if keyword:
        query = query.filter(Designer.name.contains(keyword))
    designers = query.all()
    return jsonify({'code': 200, 'data': [d.to_dict() for d in designers]})


@app.route('/api/designers', methods=['POST'])
@edit_required
def api_create_designer():
    """新增设计人员（admin/engineer）"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    name = data.get('name', '').strip()
    if not name:
        return jsonify({'code': 400, 'msg': '姓名不能为空'}), 400

    designer = Designer(
        name=name,
        department=data.get('department', '').strip(),
        major=data.get('major', '').strip(),
        title=data.get('title', '').strip(),
        phone=data.get('phone', '').strip(),
    )
    db.session.add(designer)
    db.session.commit()
    return jsonify({'code': 200, 'msg': '设计人员添加成功', 'data': designer.to_dict()})


@app.route('/api/designers/<int:designer_id>', methods=['PUT'])
@edit_required
def api_update_designer(designer_id):
    """更新设计人员（admin/engineer）"""
    designer = Designer.query.get(designer_id)
    if not designer:
        return jsonify({'code': 404, 'msg': '设计人员不存在'}), 404

    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    if 'name' in data:
        designer.name = data['name'].strip()
    if 'department' in data:
        designer.department = data['department'].strip()
    if 'major' in data:
        designer.major = data['major'].strip()
    if 'title' in data:
        designer.title = data['title'].strip()
    if 'phone' in data:
        designer.phone = data['phone'].strip()
    if 'isActive' in data:
        designer.is_active = data['isActive']

    db.session.commit()
    return jsonify({'code': 200, 'msg': '设计人员更新成功', 'data': designer.to_dict()})


@app.route('/api/designers/<int:designer_id>', methods=['DELETE'])
@admin_required
def api_delete_designer(designer_id):
    """删除设计人员（admin）"""
    designer = Designer.query.get(designer_id)
    if not designer:
        return jsonify({'code': 404, 'msg': '设计人员不存在'}), 404

    db.session.delete(designer)
    db.session.commit()
    return jsonify({'code': 200, 'msg': '设计人员已删除'})


@app.route('/api/designers/toggle-active/<int:designer_id>', methods=['POST'])
@edit_required
def api_toggle_designer_active(designer_id):
    """切换设计人员启用/禁用状态（admin/engineer）"""
    designer = Designer.query.get(designer_id)
    if not designer:
        return jsonify({'code': 404, 'msg': '设计人员不存在'}), 404

    designer.is_active = not designer.is_active
    db.session.commit()
    return jsonify({'code': 200, 'msg': '状态已更新', 'data': designer.to_dict()})


@app.route('/api/designers/export-template', methods=['GET'])
@login_required
def api_export_designer_template():
    """导出设计人员导入模板"""
    headers = ['姓名', '所属部门', '专业', '职称', '联系电话']
    
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet('设计人员模板')
    
    # 设置表头样式
    header_style = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#f0f0f0',
        'border': 1
    })
    
    # 写入表头
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_style)
    
    # 设置列宽
    worksheet.set_column(0, 0, 15)  # 姓名
    worksheet.set_column(1, 1, 20)  # 所属部门
    worksheet.set_column(2, 2, 15)  # 专业
    worksheet.set_column(3, 3, 15)  # 职称
    worksheet.set_column(4, 4, 18)  # 联系电话
    
    workbook.close()
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name='设计人员导入模板.xlsx',
        as_attachment=True
    )


@app.route('/api/designers/export-excel', methods=['GET'])
@login_required
def api_export_designers_excel():
    """导出设计人员数据到Excel"""
    designers = Designer.query.order_by(Designer.id).all()
    
    headers = ['序号', '姓名', '所属部门', '专业', '职称', '联系电话', '状态']
    
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet('设计人员')
    
    # 设置表头样式
    header_style = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#f0f0f0',
        'border': 1
    })
    
    # 设置状态样式
    active_style = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'color': '#10b981', 'border': 1})
    inactive_style = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'color': '#6b7280', 'border': 1})
    
    # 写入表头
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_style)
    
    # 写入数据
    row_num = 1
    for designer in designers:
        worksheet.write(row_num, 0, row_num)
        worksheet.write(row_num, 1, designer.name)
        worksheet.write(row_num, 2, designer.department or '')
        worksheet.write(row_num, 3, designer.major or '')
        worksheet.write(row_num, 4, designer.title or '')
        worksheet.write(row_num, 5, designer.phone or '')
        worksheet.write(row_num, 6, '启用' if designer.is_active else '禁用', active_style if designer.is_active else inactive_style)
        row_num += 1
    
    # 设置列宽
    worksheet.set_column(0, 0, 8)   # 序号
    worksheet.set_column(1, 1, 12)  # 姓名
    worksheet.set_column(2, 2, 18)  # 所属部门
    worksheet.set_column(3, 3, 12)  # 专业
    worksheet.set_column(4, 4, 12)  # 职称
    worksheet.set_column(5, 5, 16)  # 联系电话
    worksheet.set_column(6, 6, 10)  # 状态
    
    workbook.close()
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name='设计人员数据.xlsx',
        as_attachment=True
    )


@app.route('/api/designers/import-excel', methods=['POST'])
@edit_required
def api_import_designers_excel():
    """从Excel导入设计人员数据"""
    if 'file' not in request.files:
        return jsonify({'code': 400, 'msg': '请选择要导入的文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'code': 400, 'msg': '请选择要导入的文件'}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'code': 400, 'msg': '仅支持Excel文件（.xlsx/.xls）'}), 400
    
    try:
        # 读取Excel文件
        if file.filename.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file)
        else:
            workbook = xlrd.open_workbook(file_contents=file.read())
        
        worksheet = workbook.active if hasattr(workbook, 'active') else workbook.sheet_by_index(0)
        
        # 获取所有现有设计人员用于重复检测
        existing_designers = Designer.query.all()
        existing_keys = {}
        for d in existing_designers:
            key = f"{d.name}_{d.major}_{d.department}_{d.title}"
            existing_keys[key] = d
        
        imported_count = 0
        updated_count = 0
        
        # 从第二行开始读取（跳过表头）
        start_row = 1
        for row_idx in range(start_row, worksheet.max_row if hasattr(worksheet, 'max_row') else worksheet.nrows):
            try:
                # 获取单元格值
                if hasattr(worksheet, 'cell'):
                    name = str(worksheet.cell(row=row_idx+1, column=1).value).strip() if worksheet.cell(row=row_idx+1, column=1).value else ''
                    department = str(worksheet.cell(row=row_idx+1, column=2).value).strip() if worksheet.cell(row=row_idx+1, column=2).value else ''
                    major = str(worksheet.cell(row=row_idx+1, column=3).value).strip() if worksheet.cell(row=row_idx+1, column=3).value else ''
                    title = str(worksheet.cell(row=row_idx+1, column=4).value).strip() if worksheet.cell(row=row_idx+1, column=4).value else ''
                    phone = str(worksheet.cell(row=row_idx+1, column=5).value).strip() if worksheet.cell(row=row_idx+1, column=5).value else ''
                else:
                    name = str(worksheet.cell(row_idx, 0).value).strip() if worksheet.cell(row_idx, 0).value else ''
                    department = str(worksheet.cell(row_idx, 1).value).strip() if worksheet.cell(row_idx, 1).value else ''
                    major = str(worksheet.cell(row_idx, 2).value).strip() if worksheet.cell(row_idx, 2).value else ''
                    title = str(worksheet.cell(row_idx, 3).value).strip() if worksheet.cell(row_idx, 3).value else ''
                    phone = str(worksheet.cell(row_idx, 4).value).strip() if worksheet.cell(row_idx, 4).value else ''
                
                # 跳过空行
                if not name:
                    continue
                
                # 检查重复
                key = f"{name}_{major}_{department}_{title}"
                
                if key in existing_keys:
                    # 重复数据，覆盖更新
                    existing_designer = existing_keys[key]
                    existing_designer.department = department
                    existing_designer.major = major
                    existing_designer.title = title
                    existing_designer.phone = phone
                    existing_designer.is_active = True
                    updated_count += 1
                else:
                    # 新数据
                    new_designer = Designer(
                        name=name,
                        department=department,
                        major=major,
                        title=title,
                        phone=phone,
                        is_active=True
                    )
                    db.session.add(new_designer)
                    imported_count += 1
            
            except Exception as e:
                continue
        
        db.session.commit()
        
        return jsonify({
            'code': 200,
            'msg': f'导入完成！新增 {imported_count} 人，更新 {updated_count} 人'
        })
    
    except Exception as e:
        return jsonify({'code': 500, 'msg': f'导入失败：{str(e)}'}), 500


@app.route('/api/designers/suggestions', methods=['GET'])
@login_required
def api_designer_suggestions():
    """联想搜索（根据首字匹配name，返回前10条）"""
    keyword = request.args.get('keyword', '').strip()
    if not keyword:
        return jsonify({'code': 200, 'data': []})

    designers = Designer.query.filter(
        Designer.is_active == True,
        Designer.name.startswith(keyword)
    ).order_by(Designer.name).limit(10).all()

    return jsonify({'code': 200, 'data': [d.to_dict() for d in designers]})


# ==================== Major Category API ====================

@app.route('/api/major-categories', methods=['GET'])
@login_required
def api_get_major_categories():
    """获取专业类别列表"""
    categories = MajorCategory.query.order_by(MajorCategory.id).all()
    return jsonify({'code': 200, 'data': [c.to_dict() for c in categories]})


@app.route('/api/major-categories', methods=['POST'])
@edit_required
def api_create_major_category():
    """新增专业类别（admin/engineer）"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    name = data.get('name', '').strip()
    if not name:
        return jsonify({'code': 400, 'msg': '专业类别名称不能为空'}), 400

    if MajorCategory.query.filter_by(name=name).first():
        return jsonify({'code': 400, 'msg': '该专业类别已存在'}), 400

    category = MajorCategory(name=name)
    db.session.add(category)
    db.session.commit()
    return jsonify({'code': 200, 'msg': '专业类别添加成功', 'data': category.to_dict()})


@app.route('/api/major-categories/<int:category_id>', methods=['DELETE'])
@edit_required
def api_delete_major_category(category_id):
    """删除专业类别（admin/engineer）"""
    category = MajorCategory.query.get(category_id)
    if not category:
        return jsonify({'code': 404, 'msg': '专业类别不存在'}), 404

    category_name = category.name
    db.session.delete(category)
    db.session.commit()
    return jsonify({'code': 200, 'msg': '专业类别已删除'})


# ==================== Construction Plan API ====================

@app.route('/api/construction', methods=['GET'])
@login_required
def api_get_construction():
    """获取施工图计划列表"""
    keyword = request.args.get('keyword', '').strip()
    query = ConstructionPlan.query.order_by(ConstructionPlan.id)
    if keyword:
        query = query.filter(db.or_(
            ConstructionPlan.project_name.contains(keyword),
            ConstructionPlan.unit_project.contains(keyword),
            ConstructionPlan.drawing_content.contains(keyword),
            ConstructionPlan.major_category.contains(keyword),
            ConstructionPlan.designer.contains(keyword),
            ConstructionPlan.project_manager.contains(keyword),
        ))
    records = query.all()
    return jsonify({'code': 200, 'data': [r.to_dict() for r in records]})


@app.route('/api/construction', methods=['POST'])
@edit_required
def api_create_construction():
    """新增施工图计划"""
    data = request.get_json()
    if not data or not data.get('projectName'):
        return jsonify({'code': 400, 'msg': '项目名称不能为空'}), 400

    record = ConstructionPlan(
        project_name=data.get('projectName', ''),
        unit_project=data.get('unitProject', ''),
        drawing_content=data.get('drawingContent', ''),
        major_category=data.get('majorCategory', ''),
        designer=data.get('designer', ''),
        completion_status=data.get('completionStatus', ''),
        start_date=data.get('startDate', ''),
        end_date=data.get('endDate', ''),
        project_manager=data.get('projectManager', ''),
        remarks=data.get('remarks', ''),
        created_by=session.get('user_id')
    )
    db.session.add(record)
    db.session.commit()
    return jsonify({'code': 200, 'msg': '添加成功', 'data': record.to_dict()})


@app.route('/api/construction/<int:record_id>', methods=['PUT'])
@edit_required
def api_update_construction(record_id):
    """更新施工图计划"""
    record = ConstructionPlan.query.get(record_id)
    if not record:
        return jsonify({'code': 404, 'msg': '记录不存在'}), 404

    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    record.project_name = data.get('projectName', record.project_name)
    record.unit_project = data.get('unitProject', record.unit_project)
    record.drawing_content = data.get('drawingContent', record.drawing_content)
    record.major_category = data.get('majorCategory', record.major_category)
    record.designer = data.get('designer', record.designer)
    record.completion_status = data.get('completionStatus', record.completion_status)
    record.start_date = data.get('startDate', record.start_date)
    record.end_date = data.get('endDate', record.end_date)
    record.project_manager = data.get('projectManager', record.project_manager)
    record.remarks = data.get('remarks', record.remarks)

    db.session.commit()
    return jsonify({'code': 200, 'msg': '更新成功', 'data': record.to_dict()})


@app.route('/api/construction/<int:record_id>', methods=['DELETE'])
@edit_required
def api_delete_construction(record_id):
    """删除施工图计划"""
    record = ConstructionPlan.query.get(record_id)
    if not record:
        return jsonify({'code': 404, 'msg': '记录不存在'}), 404
    db.session.delete(record)
    db.session.commit()
    return jsonify({'code': 200, 'msg': '删除成功'})


# ==================== Phase Plan API ====================

@app.route('/api/phase', methods=['GET'])
@login_required
def api_get_phase():
    """获取阶段设计计划列表"""
    keyword = request.args.get('keyword', '').strip()
    query = PhasePlan.query.order_by(PhasePlan.id)
    if keyword:
        query = query.filter(db.or_(
            PhasePlan.project_name.contains(keyword),
            PhasePlan.design_phase.contains(keyword),
            PhasePlan.project_manager.contains(keyword),
            PhasePlan.department.contains(keyword),
            PhasePlan.participating_majors.contains(keyword),
        ))
    records = query.all()
    return jsonify({'code': 200, 'data': [r.to_dict() for r in records]})


@app.route('/api/phase', methods=['POST'])
@edit_required
def api_create_phase():
    """新增阶段设计计划"""
    data = request.get_json()
    if not data or not data.get('projectName'):
        return jsonify({'code': 400, 'msg': '工程名称不能为空'}), 400

    record = PhasePlan(
        project_name=data.get('projectName', ''),
        design_phase=data.get('designPhase', ''),
        project_manager=data.get('projectManager', ''),
        department=data.get('department', ''),
        participating_majors=data.get('participatingMajors', ''),
        data_submit_date=data.get('dataSubmitDate', ''),
        desc_submit_date=data.get('descSubmitDate', ''),
        publish_date=data.get('publishDate', ''),
        remarks=data.get('remarks', ''),
        created_by=session.get('user_id')
    )
    db.session.add(record)
    db.session.commit()
    return jsonify({'code': 200, 'msg': '添加成功', 'data': record.to_dict()})


@app.route('/api/phase/<int:record_id>', methods=['PUT'])
@edit_required
def api_update_phase(record_id):
    """更新阶段设计计划"""
    record = PhasePlan.query.get(record_id)
    if not record:
        return jsonify({'code': 404, 'msg': '记录不存在'}), 404

    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    record.project_name = data.get('projectName', record.project_name)
    record.design_phase = data.get('designPhase', record.design_phase)
    record.project_manager = data.get('projectManager', record.project_manager)
    record.department = data.get('department', record.department)
    record.participating_majors = data.get('participatingMajors', record.participating_majors)
    record.data_submit_date = data.get('dataSubmitDate', record.data_submit_date)
    record.desc_submit_date = data.get('descSubmitDate', record.desc_submit_date)
    record.publish_date = data.get('publishDate', record.publish_date)
    record.remarks = data.get('remarks', record.remarks)

    db.session.commit()
    return jsonify({'code': 200, 'msg': '更新成功', 'data': record.to_dict()})


@app.route('/api/phase/<int:record_id>', methods=['DELETE'])
@edit_required
def api_delete_phase(record_id):
    """删除阶段设计计划"""
    record = PhasePlan.query.get(record_id)
    if not record:
        return jsonify({'code': 404, 'msg': '记录不存在'}), 404
    db.session.delete(record)
    db.session.commit()
    return jsonify({'code': 200, 'msg': '删除成功'})


# ==================== Technical Plan API ====================

@app.route('/api/technical', methods=['GET'])
@login_required
def api_get_technical():
    """获取技术要求计划列表"""
    keyword = request.args.get('keyword', '').strip()
    query = TechnicalPlan.query.order_by(TechnicalPlan.id)
    if keyword:
        query = query.filter(db.or_(
            TechnicalPlan.equipment_name.contains(keyword),
            TechnicalPlan.designer.contains(keyword),
        ))
    records = query.all()
    return jsonify({'code': 200, 'data': [r.to_dict() for r in records]})


@app.route('/api/technical', methods=['POST'])
@edit_required
def api_create_technical():
    """新增技术要求计划"""
    data = request.get_json()
    if not data or not data.get('equipmentName'):
        return jsonify({'code': 400, 'msg': '设备或货物名称不能为空'}), 400

    record = TechnicalPlan(
        equipment_name=data.get('equipmentName', ''),
        designer=data.get('designer', ''),
        design_complete_date=data.get('designCompleteDate', ''),
        purchase_date=data.get('purchaseDate', ''),
        owner_review_date=data.get('ownerReviewDate', ''),
        owner_feedback_date=data.get('ownerFeedbackDate', ''),
        created_by=session.get('user_id')
    )
    db.session.add(record)
    db.session.commit()
    return jsonify({'code': 200, 'msg': '添加成功', 'data': record.to_dict()})


@app.route('/api/technical/<int:record_id>', methods=['PUT'])
@edit_required
def api_update_technical(record_id):
    """更新技术要求计划"""
    record = TechnicalPlan.query.get(record_id)
    if not record:
        return jsonify({'code': 404, 'msg': '记录不存在'}), 404

    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    record.equipment_name = data.get('equipmentName', record.equipment_name)
    record.designer = data.get('designer', record.designer)
    record.design_complete_date = data.get('designCompleteDate', record.design_complete_date)
    record.purchase_date = data.get('purchaseDate', record.purchase_date)
    record.owner_review_date = data.get('ownerReviewDate', record.owner_review_date)
    record.owner_feedback_date = data.get('ownerFeedbackDate', record.owner_feedback_date)

    db.session.commit()
    return jsonify({'code': 200, 'msg': '更新成功', 'data': record.to_dict()})


@app.route('/api/technical/<int:record_id>', methods=['DELETE'])
@edit_required
def api_delete_technical(record_id):
    """删除技术要求计划"""
    record = TechnicalPlan.query.get(record_id)
    if not record:
        return jsonify({'code': 404, 'msg': '记录不存在'}), 404
    db.session.delete(record)
    db.session.commit()
    return jsonify({'code': 200, 'msg': '删除成功'})


# ==================== Batch Delete API ====================

@app.route('/api/batch-delete/construction', methods=['POST'])
@edit_required
def api_batch_delete_construction():
    """批量删除施工图计划"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    query = ConstructionPlan.query

    if 'project' in data and data['project']:
        query = query.filter(ConstructionPlan.project_name == data['project'])
    if 'unit_project' in data and data['unit_project']:
        query = query.filter(ConstructionPlan.unit_project == data['unit_project'])
    if 'designer' in data and data['designer']:
        query = query.filter(ConstructionPlan.designer == data['designer'])
    if 'major' in data and data['major']:
        query = query.filter(ConstructionPlan.major_category == data['major'])
    if 'start_date' in data and data['start_date']:
        query = query.filter(ConstructionPlan.end_date >= data['start_date'])
    if 'end_date' in data and data['end_date']:
        query = query.filter(ConstructionPlan.end_date <= data['end_date'])

    records = query.all()
    count = len(records)
    
    for record in records:
        db.session.delete(record)
    db.session.commit()

    return jsonify({'code': 200, 'msg': '批量删除成功', 'data': {'count': count}})


@app.route('/api/batch-delete/phase', methods=['POST'])
@edit_required
def api_batch_delete_phase():
    """批量删除阶段设计计划"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    query = PhasePlan.query

    if 'project' in data and data['project']:
        query = query.filter(PhasePlan.project_name == data['project'])
    if 'department' in data and data['department']:
        query = query.filter(PhasePlan.department == data['department'])
    if 'start_date' in data and data['start_date']:
        query = query.filter(PhasePlan.publish_date >= data['start_date'])
    if 'end_date' in data and data['end_date']:
        query = query.filter(PhasePlan.publish_date <= data['end_date'])

    records = query.all()
    count = len(records)
    
    for record in records:
        db.session.delete(record)
    db.session.commit()

    return jsonify({'code': 200, 'msg': '批量删除成功', 'data': {'count': count}})


@app.route('/api/batch-delete/technical', methods=['POST'])
@edit_required
def api_batch_delete_technical():
    """批量删除技术要求计划"""
    data = request.get_json()
    if not data:
        return jsonify({'code': 400, 'msg': '请求数据无效'}), 400

    query = TechnicalPlan.query

    if 'equipment' in data and data['equipment']:
        query = query.filter(TechnicalPlan.equipment_name.contains(data['equipment']))
    if 'designer' in data and data['designer']:
        query = query.filter(TechnicalPlan.designer == data['designer'])
    if 'start_date' in data and data['start_date']:
        query = query.filter(TechnicalPlan.design_complete_date >= data['start_date'])
    if 'end_date' in data and data['end_date']:
        query = query.filter(TechnicalPlan.design_complete_date <= data['end_date'])

    records = query.all()
    count = len(records)
    
    for record in records:
        db.session.delete(record)
    db.session.commit()

    return jsonify({'code': 200, 'msg': '批量删除成功', 'data': {'count': count}})


# ==================== Excel Template / Import / Export API ====================

# 表头映射定义
EXCEL_HEADERS = {
    'construction': [
        '项目名称', '单位工程名称', '单位工程下专业图纸内容', '专业类别',
        '设计人', '完成情况', '计划开始时间', '计划完成时间', '项目负责人', '备注信息'
    ],
    'phase': [
        '工程名称', '设计阶段', '项目负责人', '责任部门',
        '参与专业', '提资时间', '说明提交时间', '最终出版时间', '备注'
    ],
    'technical': [
        '设备或货物名称', '设计人', '设计完成时间', '提采购时间',
        '提业主审查时间', '业主反馈时间'
    ],
}

EXCEL_FILENAMES = {
    'construction': '施工图计划模板.xlsx',
    'phase': '阶段设计计划模板.xlsx',
    'technical': '技术要求计划模板.xlsx',
}


@app.route('/api/template/<table_type>', methods=['GET'])
@login_required
def api_download_template(table_type):
    """下载Excel模板"""
    if table_type not in EXCEL_HEADERS:
        return jsonify({'code': 400, 'msg': '无效的表格类型，支持: construction / phase / technical'}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = table_type
    headers = EXCEL_HEADERS[table_type]
    ws.append(headers)

    # 设置表头样式
    from openpyxl.styles import Font, PatternFill, Alignment
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # 自动调整列宽
    for col_idx, header in enumerate(headers, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = max(len(header) * 2, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = EXCEL_FILENAMES.get(table_type, f'{table_type}_template.xlsx')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/api/import/<table_type>', methods=['POST'])
@edit_required
def api_import_excel(table_type):
    """导入Excel数据"""
    if table_type not in EXCEL_HEADERS:
        return jsonify({'code': 400, 'msg': '无效的表格类型，支持: construction / phase / technical'}), 400

    data = request.form
    overwrite = data.get('overwrite', 'false').lower() == 'true'

    if 'file' not in request.files:
        return jsonify({'code': 400, 'msg': '请上传文件（字段名: file）'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'code': 400, 'msg': '文件名为空'}), 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({'code': 400, 'msg': '仅支持 .xlsx 格式的Excel文件'}), 400

    try:
        wb = load_workbook(file, data_only=True)
    except Exception as e:
        return jsonify({'code': 400, 'msg': f'Excel文件解析失败: {str(e)}'}), 400

    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))

    if len(rows) < 2:
        return jsonify({'code': 400, 'msg': 'Excel文件没有数据行（至少需要表头+1行数据）'}), 400

    # 验证表头
    header_row = [str(cell).strip() if cell else '' for cell in rows[0]]
    expected_headers = EXCEL_HEADERS[table_type]
    if header_row != expected_headers:
        return jsonify({
            'code': 400,
            'msg': f'表头不匹配。期望: {expected_headers}，实际: {header_row}'
        }), 400

    success_count = 0
    updated_count = 0
    fail_count = 0
    fail_rows = []
    duplicates = []

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            values = [str(cell).strip() if cell is not None else '' for cell in row]

            if table_type == 'construction':
                if not values[0]:
                    fail_count += 1
                    fail_rows.append(row_idx)
                    continue
                
                # 检查是否存在重复记录
                existing = ConstructionPlan.query.filter(
                    ConstructionPlan.project_name == values[0],
                    ConstructionPlan.unit_project == values[1],
                    ConstructionPlan.drawing_content == values[2],
                    ConstructionPlan.major_category == values[3],
                    ConstructionPlan.designer == values[4]
                ).first()
                
                if existing:
                    if overwrite:
                        # 覆盖更新
                        existing.unit_project = values[1]
                        existing.drawing_content = values[2]
                        existing.major_category = values[3]
                        existing.designer = values[4]
                        existing.completion_status = values[5]
                        existing.start_date = values[6]
                        existing.end_date = values[7]
                        existing.project_manager = values[8]
                        existing.remarks = values[9]
                        existing.updated_at = datetime.now()
                        updated_count += 1
                    else:
                        duplicates.append({
                            'row': row_idx,
                            'project': values[0],
                            'unit': values[1],
                            'drawing': values[2],
                            'major': values[3],
                            'designer': values[4]
                        })
                        continue
                else:
                    record = ConstructionPlan(
                        project_name=values[0],
                        unit_project=values[1],
                        drawing_content=values[2],
                        major_category=values[3],
                        designer=values[4],
                        completion_status=values[5],
                        start_date=values[6],
                        end_date=values[7],
                        project_manager=values[8],
                        remarks=values[9],
                        created_by=session.get('user_id')
                    )
                    db.session.add(record)
                    success_count += 1
                    
            elif table_type == 'phase':
                if not values[0]:
                    fail_count += 1
                    fail_rows.append(row_idx)
                    continue
                record = PhasePlan(
                    project_name=values[0],
                    design_phase=values[1],
                    project_manager=values[2],
                    department=values[3],
                    participating_majors=values[4],
                    data_submit_date=values[5],
                    desc_submit_date=values[6],
                    publish_date=values[7],
                    remarks=values[8],
                    created_by=session.get('user_id')
                )
                db.session.add(record)
                success_count += 1
                
            elif table_type == 'technical':
                if not values[0]:
                    fail_count += 1
                    fail_rows.append(row_idx)
                    continue
                record = TechnicalPlan(
                    equipment_name=values[0],
                    designer=values[1],
                    design_complete_date=values[2],
                    purchase_date=values[3],
                    owner_review_date=values[4],
                    owner_feedback_date=values[5],
                    created_by=session.get('user_id')
                )
                db.session.add(record)
                success_count += 1
                
            else:
                fail_count += 1
                fail_rows.append(row_idx)
                continue

        except Exception:
            fail_count += 1
            fail_rows.append(row_idx)

    # 如果发现重复且未选择覆盖，则返回需要确认
    if duplicates and not overwrite:
        return jsonify({
            'code': 300,
            'msg': f'发现 {len(duplicates)} 条重复记录',
            'data': {
                'duplicates': duplicates,
                'success': success_count,
                'fail': fail_count,
                'failRows': fail_rows
            }
        })

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'code': 500, 'msg': f'数据库提交失败: {str(e)}'}), 500

    result_msg = f'导入完成：新增 {success_count} 条'
    if updated_count > 0:
        result_msg += f'，更新 {updated_count} 条'
    if fail_count > 0:
        result_msg += f'，失败 {fail_count} 条（行号: {fail_rows}）'

    return jsonify({
        'code': 200,
        'msg': result_msg,
        'data': {
            'success': success_count,
            'updated': updated_count,
            'fail': fail_count,
            'failRows': fail_rows
        }
    })


@app.route('/api/export-excel/<table_type>', methods=['GET'])
@login_required
def api_export_excel(table_type):
    """导出Excel"""
    if table_type not in EXCEL_HEADERS:
        return jsonify({'code': 400, 'msg': '无效的表格类型，支持: construction / phase / technical'}), 400

    keyword = request.args.get('keyword', '').strip()
    project = request.args.get('project', '').strip()
    major = request.args.get('major', '').strip()

    # 根据类型查询数据
    if table_type == 'construction':
        query = ConstructionPlan.query.order_by(ConstructionPlan.id)
        if keyword:
            query = query.filter(db.or_(
                ConstructionPlan.project_name.contains(keyword),
                ConstructionPlan.unit_project.contains(keyword),
                ConstructionPlan.drawing_content.contains(keyword),
                ConstructionPlan.major_category.contains(keyword),
                ConstructionPlan.designer.contains(keyword),
                ConstructionPlan.project_manager.contains(keyword),
            ))
        if project:
            query = query.filter(ConstructionPlan.project_name == project)
        if major:
            query = query.filter(ConstructionPlan.major_category == major)
        records = query.all()
        headers = EXCEL_HEADERS['construction']
        data_rows = [[
            r.project_name, r.unit_project, r.drawing_content, r.major_category,
            r.designer, r.completion_status, r.start_date, r.end_date,
            r.project_manager, r.remarks
        ] for r in records]

    elif table_type == 'phase':
        query = PhasePlan.query.order_by(PhasePlan.id)
        if keyword:
            query = query.filter(db.or_(
                PhasePlan.project_name.contains(keyword),
                PhasePlan.design_phase.contains(keyword),
                PhasePlan.project_manager.contains(keyword),
                PhasePlan.department.contains(keyword),
                PhasePlan.participating_majors.contains(keyword),
            ))
        if project:
            query = query.filter(PhasePlan.project_name == project)
        records = query.all()
        headers = EXCEL_HEADERS['phase']
        data_rows = [[
            r.project_name, r.design_phase, r.project_manager, r.department,
            r.participating_majors, r.data_submit_date, r.desc_submit_date,
            r.publish_date, r.remarks
        ] for r in records]

    elif table_type == 'technical':
        query = TechnicalPlan.query.order_by(TechnicalPlan.id)
        if keyword:
            query = query.filter(db.or_(
                TechnicalPlan.equipment_name.contains(keyword),
                TechnicalPlan.designer.contains(keyword),
            ))
        records = query.all()
        headers = EXCEL_HEADERS['technical']
        data_rows = [[
            r.equipment_name, r.designer, r.design_complete_date,
            r.purchase_date, r.owner_review_date, r.owner_feedback_date
        ] for r in records]
    else:
        data_rows = []
        headers = []

    wb = Workbook()
    ws = wb.active
    ws.title = table_type

    # 写入表头
    ws.append(headers)
    from openpyxl.styles import Font, PatternFill, Alignment
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # 写入数据
    for row_data in data_rows:
        ws.append(row_data)

    # 自动调整列宽
    for col_idx in range(1, len(headers) + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = len(str(headers[col_idx - 1])) * 2
        for row in data_rows:
            if col_idx <= len(row) and row[col_idx - 1]:
                max_len = max(max_len, len(str(row[col_idx - 1])) * 1.5)
        ws.column_dimensions[col_letter].width = min(max(max_len, 12), 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    export_filenames = {
        'construction': f'施工图计划_{timestamp}.xlsx',
        'phase': f'阶段设计计划_{timestamp}.xlsx',
        'technical': f'技术要求计划_{timestamp}.xlsx',
    }
    filename = export_filenames.get(table_type, f'{table_type}_{timestamp}.xlsx')

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ==================== Statistics API ====================

@app.route('/api/stats/overview', methods=['GET'])
@login_required
def api_stats_overview():
    """总览统计"""
    keyword = request.args.get('keyword', '').strip()

    # 施工图统计
    cq = ConstructionPlan.query
    if keyword:
        cq = cq.filter(db.or_(
            ConstructionPlan.project_name.contains(keyword),
            ConstructionPlan.unit_project.contains(keyword),
            ConstructionPlan.designer.contains(keyword),
        ))
    construction_total = cq.count()
    construction_completed = cq.filter(ConstructionPlan.completion_status == '已完成').count()
    construction_ongoing = cq.filter(ConstructionPlan.completion_status == '进行中').count()
    construction_not_started = cq.filter(ConstructionPlan.completion_status == '未开始').count()
    construction_delayed = cq.filter(ConstructionPlan.completion_status == '已延期').count()

    # 阶段设计统计
    pq = PhasePlan.query
    if keyword:
        pq = pq.filter(db.or_(
            PhasePlan.project_name.contains(keyword),
            PhasePlan.design_phase.contains(keyword),
        ))
    phase_total = pq.count()

    # 技术要求统计
    tq = TechnicalPlan.query
    if keyword:
        tq = tq.filter(db.or_(
            TechnicalPlan.equipment_name.contains(keyword),
            TechnicalPlan.designer.contains(keyword),
        ))
    technical_total = tq.count()

    # 专业分布（从施工图表统计）
    major_stats = {}
    cq_all = ConstructionPlan.query
    if keyword:
        cq_all = cq_all.filter(db.or_(
            ConstructionPlan.project_name.contains(keyword),
            ConstructionPlan.unit_project.contains(keyword),
            ConstructionPlan.designer.contains(keyword),
        ))
    for r in cq_all.all():
        major = r.major_category or '未分类'
        if major not in major_stats:
            major_stats[major] = 0
        major_stats[major] += 1

    return jsonify({
        'code': 200,
        'data': {
            'tables': {
                'construction': construction_total,
                'phase': phase_total,
                'technical': technical_total,
            },
            'completionStatus': {
                'labels': ['已完成', '进行中', '未开始', '已延期'],
                'values': [construction_completed, construction_ongoing, construction_not_started, construction_delayed]
            },
            'majorDistribution': {
                'labels': list(major_stats.keys()),
                'values': list(major_stats.values())
            }
        }
    })


@app.route('/api/stats/by-project', methods=['GET'])
@login_required
def api_stats_by_project():
    """按项目统计（支持过滤）"""
    project_param = request.args.get('project', '').strip()
    unit_project_param = request.args.get('unit_project', '').strip()
    major_param = request.args.get('major', '').strip()

    cq = ConstructionPlan.query

    # 按项目过滤（支持逗号分隔多个）
    if project_param:
        project_list = [p.strip() for p in project_param.split(',') if p.strip()]
        if project_list:
            cq = cq.filter(ConstructionPlan.project_name.in_(project_list))

    # 按单位工程过滤
    if unit_project_param:
        cq = cq.filter(ConstructionPlan.unit_project == unit_project_param)

    # 按专业过滤
    if major_param:
        major_list = [m.strip() for m in major_param.split(',') if m.strip()]
        if major_list:
            cq = cq.filter(ConstructionPlan.major_category.in_(major_list))

    records = cq.all()

    project_stats = {}
    for r in records:
        pname = r.project_name or '未命名项目'
        if pname not in project_stats:
            project_stats[pname] = {'total': 0, 'completed': 0, 'ongoing': 0, 'notStarted': 0, 'delayed': 0, 'items': []}
        project_stats[pname]['total'] += 1
        status = r.completion_status or ''
        if status == '已完成':
            project_stats[pname]['completed'] += 1
        elif status == '进行中':
            project_stats[pname]['ongoing'] += 1
        elif status == '未开始':
            project_stats[pname]['notStarted'] += 1
        elif status == '已延期':
            project_stats[pname]['delayed'] += 1
        project_stats[pname]['items'].append({
            'unitProject': r.unit_project or '',
            'drawingContent': r.drawing_content or '',
            'major': r.major_category or '',
            'designer': r.designer or '',
            'completionStatus': r.completion_status or '',
        })

    labels = list(project_stats.keys())
    details = []
    for label in labels:
        s = project_stats[label]
        details.append({
            'projectName': label,
            'total': s['total'],
            'completed': s['completed'],
            'ongoing': s['ongoing'],
            'notStarted': s['notStarted'],
            'delayed': s['delayed'],
            'items': s['items'],
        })

    return jsonify({
        'code': 200,
        'data': {
            'labels': labels,
            'totals': [project_stats[l]['total'] for l in labels],
            'completed': [project_stats[l]['completed'] for l in labels],
            'details': details,
        }
    })


@app.route('/api/stats/by-unit', methods=['GET'])
@login_required
def api_stats_by_unit():
    """按单位工程统计"""
    keyword = request.args.get('keyword', '').strip()

    cq = ConstructionPlan.query
    if keyword:
        cq = cq.filter(db.or_(
            ConstructionPlan.project_name.contains(keyword),
            ConstructionPlan.unit_project.contains(keyword),
            ConstructionPlan.designer.contains(keyword),
        ))
    records = cq.all()

    unit_stats = {}
    for r in records:
        uname = r.unit_project or '未分类'
        if uname not in unit_stats:
            unit_stats[uname] = 0
        unit_stats[uname] += 1

    # 按数量降序排列
    sorted_units = sorted(unit_stats.items(), key=lambda x: x[1], reverse=True)

    return jsonify({
        'code': 200,
        'data': {
            'labels': [u[0] for u in sorted_units],
            'values': [u[1] for u in sorted_units]
        }
    })


@app.route('/api/stats/by-major', methods=['GET'])
@login_required
def api_stats_by_major():
    """按专业统计（支持过滤）"""
    project = request.args.get('project', '').strip()
    start_date_str = request.args.get('start_date', '').strip()
    end_date_str = request.args.get('end_date', '').strip()
    major_param = request.args.get('major', '').strip()

    cq = ConstructionPlan.query

    # 按项目过滤
    if project:
        cq = cq.filter(ConstructionPlan.project_name == project)

    # 按日期范围过滤（基于endDate）
    if start_date_str:
        cq = cq.filter(ConstructionPlan.end_date >= start_date_str)
    if end_date_str:
        cq = cq.filter(ConstructionPlan.end_date <= end_date_str)

    # 按专业过滤
    if major_param:
        major_list = [m.strip() for m in major_param.split(',') if m.strip()]
        if major_list:
            cq = cq.filter(ConstructionPlan.major_category.in_(major_list))

    records = cq.all()

    major_stats = {}
    for r in records:
        major = r.major_category or '未分类'
        if major not in major_stats:
            major_stats[major] = {'total': 0, 'completed': 0, 'ongoing': 0, 'notStarted': 0, 'delayed': 0, 'items': []}
        major_stats[major]['total'] += 1
        status = r.completion_status or ''
        if status == '已完成':
            major_stats[major]['completed'] += 1
        elif status == '进行中':
            major_stats[major]['ongoing'] += 1
        elif status == '未开始':
            major_stats[major]['notStarted'] += 1
        elif status == '已延期':
            major_stats[major]['delayed'] += 1
        major_stats[major]['items'].append({
            'projectName': r.project_name or '',
            'unitProject': r.unit_project or '',
            'drawingContent': r.drawing_content or '',
            'designer': r.designer or '',
            'completionStatus': r.completion_status or '',
            'startDate': r.start_date or '',
            'endDate': r.end_date or '',
        })

    labels = list(major_stats.keys())
    details = []
    for label in labels:
        s = major_stats[label]
        details.append({
            'major': label,
            'total': s['total'],
            'completed': s['completed'],
            'ongoing': s['ongoing'],
            'notStarted': s['notStarted'],
            'delayed': s['delayed'],
            'items': s['items'],
        })

    return jsonify({
        'code': 200,
        'data': {
            'labels': labels,
            'totals': [major_stats[l]['total'] for l in labels],
            'completed': [major_stats[l]['completed'] for l in labels],
            'ongoing': [major_stats[l]['ongoing'] for l in labels],
            'notStarted': [major_stats[l]['notStarted'] for l in labels],
            'delayed': [major_stats[l]['delayed'] for l in labels],
            'details': details,
        }
    })


@app.route('/api/stats/by-designer', methods=['GET'])
@login_required
def api_stats_by_designer():
    """按设计人员统计（支持过滤）"""
    designer_param = request.args.get('designer', '').strip()
    project = request.args.get('project', '').strip()
    start_date_str = request.args.get('start_date', '').strip()
    end_date_str = request.args.get('end_date', '').strip()

    # 施工图查询
    cq = ConstructionPlan.query
    if project:
        cq = cq.filter(ConstructionPlan.project_name == project)
    if start_date_str:
        cq = cq.filter(ConstructionPlan.end_date >= start_date_str)
    if end_date_str:
        cq = cq.filter(ConstructionPlan.end_date <= end_date_str)
    if designer_param:
        designer_list = [d.strip() for d in designer_param.split(',') if d.strip()]
        if designer_list:
            cq = cq.filter(ConstructionPlan.designer.in_(designer_list))
    construction_records = cq.all()

    # 技术要求查询
    tq = TechnicalPlan.query
    if designer_param:
        designer_list = [d.strip() for d in designer_param.split(',') if d.strip()]
        if designer_list:
            tq = tq.filter(TechnicalPlan.designer.in_(designer_list))
    technical_records = tq.all()

    designer_stats = {}
    for r in construction_records:
        dname = r.designer or '未指定'
        if dname not in designer_stats:
            designer_stats[dname] = {'total': 0, 'completed': 0, 'items': []}
        designer_stats[dname]['total'] += 1
        if r.completion_status == '已完成':
            designer_stats[dname]['completed'] += 1
        designer_stats[dname]['items'].append({
            'projectName': r.project_name or '',
            'type': '施工图',
            'content': r.drawing_content or '',
            'completionStatus': r.completion_status or '',
            'endDate': r.end_date or '',
        })

    for r in technical_records:
        dname = r.designer or '未指定'
        if dname not in designer_stats:
            designer_stats[dname] = {'total': 0, 'completed': 0, 'items': []}
        designer_stats[dname]['total'] += 1
        designer_stats[dname]['items'].append({
            'projectName': '',
            'type': '技术要求',
            'content': r.equipment_name or '',
            'completionStatus': '',
            'endDate': r.design_complete_date or '',
        })

    sorted_designers = sorted(designer_stats.items(), key=lambda x: x[1]['total'], reverse=True)
    labels = [d[0] for d in sorted_designers]
    details = []
    for dname in labels:
        s = designer_stats[dname]
        details.append({
            'designer': dname,
            'total': s['total'],
            'completed': s['completed'],
            'items': s['items'],
        })

    return jsonify({
        'code': 200,
        'data': {
            'labels': labels,
            'totals': [designer_stats[l]['total'] for l in labels],
            'completed': [designer_stats[l]['completed'] for l in labels],
            'details': details,
        }
    })


@app.route('/api/stats/by-time', methods=['GET'])
@login_required
def api_stats_by_time():
    """按时间阶段统计（默认最近12个月，支持按项目过滤）"""
    project = request.args.get('project', '').strip()

    start_str = request.args.get('start_date', '').strip()
    end_str = request.args.get('end_date', '').strip()

    # 解析时间范围
    if start_str and end_str:
        try:
            start_date = datetime.strptime(start_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_str, '%Y-%m-%d')
        except ValueError:
            return jsonify({'code': 400, 'msg': '日期格式无效，请使用 YYYY-MM-DD'}), 400
    else:
        # 默认最近12个月
        end_date = datetime.now()
        start_date = end_date - timedelta(days=365)

    # 施工图按月统计（根据end_date计划完成时间）
    cq = ConstructionPlan.query.filter(ConstructionPlan.end_date != '', ConstructionPlan.end_date.isnot(None))
    if project:
        cq = cq.filter(ConstructionPlan.project_name == project)
    construction_records = cq.all()

    # 阶段设计按月统计（根据publish_date）
    pq = PhasePlan.query.filter(PhasePlan.publish_date != '', PhasePlan.publish_date.isnot(None))
    if project:
        pq = pq.filter(PhasePlan.project_name == project)
    phase_records = pq.all()

    # 技术要求按月统计（根据design_complete_date）
    tq = TechnicalPlan.query.filter(TechnicalPlan.design_complete_date != '', TechnicalPlan.design_complete_date.isnot(None))
    technical_records = tq.all()

    # 生成月份标签
    monthly_data = {}
    current = start_date.replace(day=1)
    while current <= end_date:
        key = current.strftime('%Y-%m')
        monthly_data[key] = {'construction': 0, 'phase': 0, 'technical': 0, 'completed': 0}
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)

    for r in construction_records:
        if r.end_date:
            try:
                dt = datetime.strptime(r.end_date, '%Y-%m-%d')
                if start_date <= dt <= end_date:
                    key = dt.strftime('%Y-%m')
                    if key in monthly_data:
                        monthly_data[key]['construction'] += 1
                        if r.completion_status == '已完成':
                            monthly_data[key]['completed'] += 1
            except ValueError:
                pass

    for r in phase_records:
        if r.publish_date:
            try:
                dt = datetime.strptime(r.publish_date, '%Y-%m-%d')
                if start_date <= dt <= end_date:
                    key = dt.strftime('%Y-%m')
                    if key in monthly_data:
                        monthly_data[key]['phase'] += 1
            except ValueError:
                pass

    for r in technical_records:
        if r.design_complete_date:
            try:
                dt = datetime.strptime(r.design_complete_date, '%Y-%m-%d')
                if start_date <= dt <= end_date:
                    key = dt.strftime('%Y-%m')
                    if key in monthly_data:
                        monthly_data[key]['technical'] += 1
            except ValueError:
                pass

    labels = list(monthly_data.keys())
    totals = [monthly_data[k]['construction'] for k in labels]
    completed = [monthly_data[k]['completed'] for k in labels]
    phase_totals = [monthly_data[k]['phase'] for k in labels]
    technical_totals = [monthly_data[k]['technical'] for k in labels]

    return jsonify({
        'code': 200,
        'data': {
            'labels': labels,
            'totals': totals,
            'completed': completed,
            'phaseTotals': phase_totals,
            'technicalTotals': technical_totals
        }
    })


@app.route('/api/stats/filter-options', methods=['GET'])
@login_required
def api_stats_filter_options():
    """返回所有可用的筛选选项（用于前端下拉框）"""
    # 从施工图表提取
    construction_records = ConstructionPlan.query.all()
    # 从阶段设计表提取
    phase_records = PhasePlan.query.all()
    # 从技术要求表提取
    technical_records = TechnicalPlan.query.all()

    projects = set()
    majors = set()
    designers = set()
    unit_projects = set()
    phases = set()

    for r in construction_records:
        if r.project_name:
            projects.add(r.project_name)
        if r.major_category:
            majors.add(r.major_category)
        if r.designer:
            designers.add(r.designer)
        if r.unit_project:
            unit_projects.add(r.unit_project)

    for r in phase_records:
        if r.project_name:
            projects.add(r.project_name)
        if r.design_phase:
            phases.add(r.design_phase)

    for r in technical_records:
        if r.designer:
            designers.add(r.designer)

    return jsonify({
        'code': 200,
        'data': {
            'projects': sorted(list(projects)),
            'majors': sorted(list(majors)),
            'designers': sorted(list(designers)),
            'unitProjects': sorted(list(unit_projects)),
            'phases': sorted(list(phases)),
        }
    })


# ==================== Error Handlers ====================

@app.errorhandler(404)
def not_found(e):
    return jsonify({'code': 404, 'msg': '资源不存在'}), 404


@app.errorhandler(500)
def server_error(e):
    return jsonify({'code': 500, 'msg': '服务器内部错误'}), 500


# ==================== Main ====================

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        init_default_data()
    print('=' * 50)
    print('  设计项目进度管理系统 已启动')
    print('  访问地址: http://localhost:5000')
    print('  默认账号: admin / admin123')
    print('=' * 50)
    app.run(host='0.0.0.0', port=5000, debug=True)
